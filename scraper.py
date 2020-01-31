from openpyxl import load_workbook
from mapping import ID,NAME,EMAIL
from classes import Violation

class Scraper:
    def __init__(self, file):
        self.file = file
        self.violations = []
        self.workbook = load_workbook(self.file, read_only=True)
        self.sheet = self.active_sheet()
        
    def active_sheet(self): 
        return self.workbook.active

    def get_violations(self, min_row = 1, max_row = 2):
        for row in self.sheet.iter_rows(min_row, max_row, values_only=True):
            violation = Violation(
                id=row[ID], 
                name=row[NAME],
                email=row[EMAIL]
        ) 

        self.violations.append(violation)
        return self.violations